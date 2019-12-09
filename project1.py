from openpyxl import *
from textblob import TextBlob
from statistics import *
from wordcloud import WordCloud
from calcStarAverages import *
from calcSentiment import *
from Calculations import *
from MinCalc import *
from maxCalc import *
import matplotlib.pyplot as plt

#load the workbook and set sheet1 active
wk1 = load_workbook("AmazonTabletReviews.xlsx")
sheet1 = wk1.active
wk2 = Workbook()
sheetW = wk2.active

#setting each label to each columnheader
userNameLabel = sheet1['I1'].value
titleLabel = sheet1['H1'].value
reviewLabel = sheet1['G1'].value
starsLabel = sheet1['F1'].value
nameLabel = sheet1['B1'].value
prodIDLabel = sheet1['A1'].value

#setting each variable to the second cell in each column 
userName = sheet1.cell(column = 9, row = 2).value
reviewTitle = sheet1.cell(column = 8, row = 2).value
reviewText = sheet1.cell(column = 7, row = 2).value
starRating = sheet1.cell(column = 6, row = 2).value
prodName = sheet1.cell(column = 2, row = 2).value
prodID = sheet1.cell(column = 1, row = 2).value

#list for all stars and all reviews, and all review titles for subjectivity and polarity
starList = []
reviewList = []
titleList = []
reviewSubList = []
titleSubList = []

#list 1 for product 1 stars and reviews, and review titles for subjectivity and polarity
starList1 = []
reviewList1 = []
titleList1 = []
reviewSubList1 = []
titleSubList1 = []

#list 2 for product 2 stars and reviews, and review titles for subjectivity and polarity
starList2 = []
reviewList2 = []
titleList2 = []
reviewSubList2 = []
titleSubList2 = []

#list 3 for product 3 stars and reviews, and review titles for subjectivity and polarity
starList3 = []
reviewList3 = []
titleList3 = []
reviewSubList3 = []
titleSubList3 = []

#list 4 for product 4 stars and reviews, and review titles for subjectivity and polarity
starList4 = []
reviewList4 = []
titleList4 = []
reviewSubList4 = []
titleSubList4 = []

#list 5 for product 5 stars and reviews, and review titles for subjectivity and polarity
starList5 = []
reviewList5 = []
titleList5 = []
reviewSubList5 = []
titleSubList5 = []

#list 6 for product 6 stars and reviews, and review titles for subjectivity and polarity
starList6 = []
reviewList6 = []
titleList6 = []
reviewSubList6 = []
titleSubList6 = []

#List for the number of positive and negative reviews 
rposP = []
rnegP = []

tposP = []
tnegP = []

rposP1 = []
rnegP1 = []

tposP1 = []
tnegP1 = []

rposP2 = []
rnegP2 = []

tposP2 = []
tnegP2 = []

rposP3 = []
rnegP3 = []

tposP3 = []
tnegP3 = []

rposP4 = []
rnegP4 = []

tposP4 = []
tnegP4 = []

rposP5 = []
rnegP5 = []

tposP5 = []
tnegP5 = []

rposP6 = []
rnegP6 = []

tposP6 = []
tnegP6 = []


#AVqkIhwDv8e3D1O-lebb
#AVqkIiKWnnc1JgDc3khH
#AVqkIj9snnc1JgDc3khU
#AVqVGZNvQMlgsOJE6eUY
#AVpfwS_CLJeJML43DH5w
#AVphgVaX1cnluZ0-DR74

#set each variable to differnet product names 
prodNum = "All Amazon Tablets"
prodNum1 = "AVqkIhwDv8e3D1O-lebb"
prodNum2 = "AVqkIiKWnnc1JgDc3khH"
prodNum3 = "AVqkIj9snnc1JgDc3khU"
prodNum4 = "AVqVGZNvQMlgsOJE6eUY"
prodNum5 = "AVpfwS_CLJeJML43DH5w"
prodNum6 = "AVphgVaX1cnluZ0-DR74"

#dictionary for products. Includes product ID to their appropriate product name 
productDictionary = {}

#Word cloud for titles 
myText1 = ""
myText2 = ""

#Word cloud for review text 
#setting i = 2 so that the we increment starting at row 2
i = 2
#requirement length for titles 
reqLen = 4

#while loop to not count any rows without product ID 
while prodID != None and reviewTitle != None and reviewText != None:

    starRating = sheet1.cell(column = 6, row = i).value
    prodID = sheet1.cell(column = 1, row = i).value
    prodName = sheet1.cell(column = 2, row = i).value 
    
    #changing the data type for reviews and titles with numbers. 
    reviewTitle = str(sheet1.cell(column = 8, row = i).value)
    reviewText = str(sheet1.cell(column = 7, row = i).value)

    if reviewTitle != None and reviewText != None:

        myText1 = myText1 + " " + reviewTitle
        myText2 = myText2 + " " + reviewText

    #adds each prodID and their respected product name to the dictionary. Dictionary does add the same productID and product name to the list
    productDictionary[prodID] = prodName

    #will not calculate any rows without stars 
    if starRating != None and starRating >= 1 and starRating <= 5:

        starRating = sheet1.cell(column = 6, row = i).value
        #added rating to starList 
        starList.append(starRating)

        #checks to see if the review title is long enough to calculate sentiment 
        if len(reviewTitle) > reqLen:

            #adds the title sub and pol to their appropriate lists it also passes review title to function to calculate polarity and subjectivity

            tPol = calcTPolarity(reviewTitle)
            titleList.append(tPol)
            tSub = calcTSubjectivity(reviewTitle)
            titleSubList.append(tSub)

            if tPol > 0:
                tposP.append(tPol)
            else:
                tnegP.append(tPol)
           
        reviewText = sheet1.cell(column = 7, row = i).value
        #adds review text sub and pol to their appropriate lists it also passes review text to function to calculate polarity and subjectivity
        rPol = calcRPolarity(reviewText)
        reviewList.append(rPol)
        rSub = calcRSubjectivity(reviewText)
        reviewSubList.append(rSub)

        if rPol > 0:
            rposP.append(rPol)
        else:
            rnegP.append(rPol)

        #checks if prodID equals to the specific string and then adds rating and sentiment to their appropriate list 
        if prodID == prodNum1:

            starRating = sheet1.cell(column = 6, row = i).value
            starList1.append(starRating)
            
            if len(reviewTitle) > reqLen:

                tPol = calcTPolarity(reviewTitle)
                titleList1.append(tPol)
                tSub = calcTSubjectivity(reviewTitle)
                titleSubList1.append(tSub)
                
                if tPol > 0:
                    tposP1.append(tPol)
                else:
                    tnegP1.append(tPol)
                
            reviewText = sheet1.cell(column = 7, row = i).value
            rPol = calcRPolarity(reviewText)
            reviewList1.append(rPol)
            rSub = calcRSubjectivity(reviewText)
            reviewSubList1.append(rSub)

            if rPol > 0:
                rposP1.append(rPol)
            else:
                rnegP1.append(rPol)

        #checks if prodID equals to the specific string and then adds rating and sentiment to their appropriate list 
        elif prodID == prodNum2:

            starRating = sheet1.cell(column = 6, row = i).value
            starList2.append(starRating)

            if len(reviewTitle) > reqLen:


                tPol = calcTPolarity(reviewTitle)
                titleList2.append(tPol)
                tSub = calcTSubjectivity(reviewTitle)
                titleSubList2.append(tSub)
                if tPol > 0:
                    tposP2.append(tPol)
                else:
                    tnegP2.append(tPol)
            
            reviewText = sheet1.cell(column = 7, row = i).value
            rPol = calcRPolarity(reviewText)
            reviewList2.append(rPol)
            rSub = calcRSubjectivity(reviewText)
            reviewSubList2.append(rSub)
            
            if rPol > 0:
                rposP2.append(rPol)
            else:
                rnegP2.append(rPol)
            
        #checks if prodID equals to the specific string and then adds rating and sentiment to their appropriate list 
        elif prodID == prodNum3:

            starRating = sheet1.cell(column = 6, row = i).value
            starList3.append(starRating)

            if len(reviewTitle) > reqLen:
            
                tPol = calcTPolarity(reviewTitle)
                titleList3.append(tPol)
                tSub = calcTSubjectivity(reviewTitle)
                titleSubList3.append(tSub)
                if tPol > 0:
                    tposP3.append(tPol)
                else:
                    tnegP3.append(tPol)
                
            reviewText = sheet1.cell(column = 7, row = i).value
            rPol = calcRPolarity(reviewText)
            reviewList3.append(rPol)
            rSub = calcRSubjectivity(reviewText)
            reviewSubList3.append(rSub)
            
            if rPol > 0:
                rposP3.append(rPol)
            else:
                rnegP3.append(rPol)
            
        #checks if prodID equals to the specific string and then adds rating and sentiment to their appropriate list 
        elif prodID == prodNum4:

            starRating = sheet1.cell(column = 6, row = i).value
            starList4.append(starRating)

            if len(reviewTitle) > reqLen:
            
                tPol = calcTPolarity(reviewTitle)
                titleList4.append(tPol)
                tSub = calcTSubjectivity(reviewTitle)
                titleSubList4.append(tSub)
                if tPol > 0:
                    tposP4.append(tPol)
                else:
                    tnegP4.append(tPol)
                
            reviewText = sheet1.cell(column = 7, row = i).value
            rPol = calcRPolarity(reviewText)
            reviewList4.append(rPol)
            rSub = calcRSubjectivity(reviewText)
            reviewSubList4.append(rSub)
            
            if rPol > 0:
                rposP4.append(rPol)
            else:
                rnegP4.append(rPol)
            
        #checks if prodID equals to the specific string and then adds rating and sentiment to their appropriate list 
        elif prodID == prodNum5:

            starRating = sheet1.cell(column = 6, row = i).value
            starList5.append(starRating)

            if len(reviewTitle) > reqLen:
            
                tPol = calcTPolarity(reviewTitle)
                titleList5.append(tPol)
                tSub = calcTSubjectivity(reviewTitle)
                titleSubList5.append(tSub)
                if tPol > 0:
                    tposP5.append(tPol)
                else:
                    tnegP5.append(tPol)
                
            reviewText = sheet1.cell(column = 7, row = i).value
            rPol = calcRPolarity(reviewText)
            reviewList5.append(rPol)
            rSub = calcRSubjectivity(reviewText)
            reviewSubList5.append(rSub)
            
            if rPol > 0:
                rposP5.append(rPol)
            else:
                rnegP5.append(rPol)
            
        #checks if prodID equals to the specific string and then adds rating and sentiment to their appropriate list 
        elif prodID == prodNum6:

            starRating = sheet1.cell(column = 6, row = i).value
            starList6.append(starRating)

            if len(reviewTitle) > reqLen:
            
                tPol = calcTPolarity(reviewTitle)
                titleList6.append(tPol)
                tSub = calcTSubjectivity(reviewTitle)
                titleSubList6.append(tSub)
                if tPol > 0:
                    tposP6.append(tPol)
                else:
                    tnegP6.append(tPol)
               
            reviewText = sheet1.cell(column = 7, row = i).value
            rPol = calcRPolarity(reviewText)
            reviewList6.append(rPol)
            rSub = calcRSubjectivity(reviewText)
            reviewSubList6.append(rSub)
            
            if rPol > 0:
                rposP6.append(rPol)
            else:
                rnegP6.append(rPol)
        
    #if there are no stars in the cell then this else statement will only calculate the sentiment and insert into list
    else:

        if prodID == prodNum1:
            
            if len(reviewTitle) > reqLen:

                tPol = calcTPolarity(reviewTitle)
                titleList1.append(tPol)
                tSub = calcTSubjectivity(reviewTitle)
                titleSubList1.append(tSub)
                if tPol > 0:
                    tposP1.append(tPol)
                else:
                    tnegP1.append(tPol)
                
            reviewText = sheet1.cell(column = 7, row = i).value
            rPol = calcRPolarity(reviewText)
            reviewList1.append(rPol)
            rSub = calcRSubjectivity(reviewText)
            reviewSubList1.append(rSub)
            
            if rPol > 0:
                rposP1.append(rPol)
            else:
                rnegP1.append(rPol)


        elif prodID == prodNum2:

            if len(reviewTitle) > reqLen:

                tPol = calcTPolarity(reviewTitle)
                titleList2.append(tPol)
                tSub = calcTSubjectivity(reviewTitle)
                titleSubList2.append(tSub)
                if tPol > 0:
                    tposP2.append(tPol)
                else:
                    tnegP2.append(tPol)
               
            reviewText = sheet1.cell(column = 7, row = i).value
            rPol = calcRPolarity(reviewText)
            reviewList2.append(rPol)
            rSub = calcRSubjectivity(reviewText)
            reviewSubList2.append(rSub)
            
            if rPol > 0:
                rposP2.append(rPol)
            else:
                rnegP2.append(rPol)

        elif prodID == prodNum3:

            if len(reviewTitle) > reqLen:

                tPol = calcTPolarity(reviewTitle)
                titleList3.append(tPol)
                tSub = calcTSubjectivity(reviewTitle)
                titleSubList3.append(tSub)
                if tPol > 0:
                    tposP3.append(tPol)
                else:
                    tnegP3.append(tPol)
                
            reviewText = sheet1.cell(column = 7, row = i).value
            rPol = calcRPolarity(reviewText)
            reviewList3.append(rPol)
            rSub = calcRSubjectivity(reviewText)
            reviewSubList3.append(rSub)
            
            if rPol > 0:
                rposP3.append(rPol)
            else:
                rnegP3.append(rPol)

        elif prodID == prodNum4:

            if len(reviewTitle) > reqLen:

                tPol = calcTPolarity(reviewTitle)
                titleList4.append(tPol)
                tSub = calcTSubjectivity(reviewTitle)
                titleSubList4.append(tSub)
                if tPol > 0:
                    tposP4.append(tPol)
                else:
                    tnegP4.append(tPol)
                
            reviewText = sheet1.cell(column = 7, row = i).value
            rPol = calcRPolarity(reviewText)
            reviewList4.append(rPol)
            rSub = calcRSubjectivity(reviewText)
            reviewSubList4.append(rSub)
            
            if rPol > 0:
                rposP4.append(rPol)
            else:
                rnegP4.append(rPol)
            
        elif prodID == prodNum5:

            if len(reviewTitle) > reqLen:

                tPol = calcTPolarity(reviewTitle)
                titleList5.append(tPol)
                tSub = calcTSubjectivity(reviewTitle)
                titleSubList5.append(tSub)
                if tPol > 0:
                    tposP5.append(tPol)
                else:
                    tnegP5.append(tPol)
            
            reviewText = sheet1.cell(column = 7, row = i).value
            rPol = calcRPolarity(reviewText)
            reviewList5.append(rPol)
            rSub = calcRSubjectivity(reviewText)
            reviewSubList5.append(rSub)
            
            if rPol > 0:
                rposP3.append(rPol)
            else:
                rnegP5.append(rPol)

        elif prodID == prodNum6:

            if len(reviewTitle) > reqLen:

                tPol = calcTPolarity(reviewTitle)
                titleList6.append(tPol)
                tSub = calcTSubjectivity(reviewTitle)
                titleSubList6.append(tSub)
                if tPol > 0:
                    tposP6.append(tPol)
                else:
                    tnegP6.append(tPol)
                
            reviewText = sheet1.cell(column = 7, row = i).value
            rPol = calcRPolarity(reviewText)
            reviewList6.append(rPol)
            rSub = calcRSubjectivity(reviewText)
            reviewSubList6.append(rSub)
            
            if rPol > 0:
                rposP6.append(rPol)
            else:
                rnegP6.append(rPol)
            
    #increment position
    i = i + 1

#prints out the whole dictionary of the prodID that we have and their names 
print("Products")
for key in productDictionary:
    print(key, " ", productDictionary[key])

#Dictionary Library into excel file
sheetW['O1'].value = "Product Dictionary"
sheetW['O2'].value = "Product ID"
sheetW['O3'].value = prodNum1
sheetW['O4'].value = prodNum2
sheetW['O5'].value = prodNum3
sheetW['O6'].value = prodNum4
sheetW['O7'].value = prodNum5
sheetW['O8'].value = prodNum6
sheetW['P2'].value = "Product Description"
sheetW['P3'].value = productDictionary[prodNum1]
sheetW['P4'].value = productDictionary[prodNum2]
sheetW['P5'].value = productDictionary[prodNum3]
sheetW['P6'].value = productDictionary[prodNum4]
sheetW['P7'].value = productDictionary[prodNum5]
sheetW['P8'].value = productDictionary[prodNum6]

print("")

#sheetW['D1'].value = "Average Star Rating"

#function that passes all variables to calculate the average rating, polarity, and subjectivity for each product 
    
def averageForLists(product, stars, titlePol, titleSub, reviewPol, reviewSub):
        
    #checks if stars do not equal zero so that the average of stars can be calculated
    if stars != 0:
        #sets variable equal to function to calculate average stars 
        global avgStars
        avgStars = calcAvg(stars)
        #prints out product id
        if product in productDictionary:
            print("Product ID: ", product)
            print("Product Name: ", productDictionary[product])
        #Prints out the variable of average rating for all products
        print("# of ratings in list: ", len(stars))
        print("Average rating: ", avgStars)

        #sets variable to funtion that finds the standard deviation
        global stddev
        stddev = standardDev(stars)
        #prints out standard deviation 
        print("Standard Deviation of star ratings: ", stddev)
        
        #sets variable equal to function to find the minumum rating
        global minimumRating
        minimumRating = minData(stars)
        #prints out the minimum rating 
        print("Minimum star rating: ", minimumRating)
        
        #sets variable equal to function to find the maximum rating
        global maximumRating
        maximumRating = maxData(stars)
        #prints out the maximum rating
        print("Maximum star rating: ", maximumRating)
        
        #sets variable equal to function to find the range of rating 
        global rangeRating
        rangeRating = rangeData(stars)
        #prints out range of star ratings
        print("Range of star rating: ", rangeRating)
        
    
        
    #loop for if star list has no rating 
    else:
        if product in productDictionary:
            print("Product ID: ", product)
            print("Product Name: ", productDictionary[product])
        print("# of ratings in list: ", 0)
        print("Average rating: ", stars)
            
    #sets variable to funtion that calculate average polarity of tile and prints it out
    global avgTitlePol
    avgTitlePol = calcAvg(titlePol)
    print("Average review title polarity: ", avgTitlePol)

    #sets variable to funtion that finds the standard deviation of Review Title Polarity
    global stddevTPol
    stddevTPol = standardDev(titlePol)
    #prints out standard deviation 
    print("Standard Deviation of review title Polarity: ", stddevTPol)
    
    #sets variable to function that finds minimum of title polarity and prints it out
    global minTPol
    minTPol = minData(titlePol)
    print("Minimum review title polarity: ", minTPol)
    
    #sets variable to function that finds maximum of title polarity and prints it out 
    global maxTPol
    maxTPol = maxData(titlePol)
    print("Maximum review title polarity: ", maxTPol)
    
    #sets variable to function that finds the range of title polarity and prints it out 
    global rangeTPol
    rangeTPol = rangeData(titlePol)
    print("Range of review title polairty: ", rangeTPol)

    #sets variable to funtion that calculate average subjectivity of title and prints it out
    global avgTitleSub
    avgTitleSub = calcAvg(titleSub)
    print("Average review title subjectivity: ", avgTitleSub)

    #sets variable to funtion that finds the standard deviation of Review Title Subjectivity
    global stddevTSub
    stddevTSub = standardDev(titleSub)
    #prints out standard deviation 
    print("Standard Deviation of review title subjectivity: ", stddevTSub)

    #sets variable to function that finds minimum subjectivity of title and prints it out
    global minTSub
    minTSub = minData(titleSub)
    print("Minimum review title subjectivity: ", minTSub)

    #sets variable to funtion that finds maximum subjectivity of title and prints it out
    global maxTSub
    maxTSub = maxData(titleSub)
    print("Maximum review title subjectivity: ", maxTSub)

    #sets variable to function that finds the range of title subjectivity and prints it out
    global rangeTSub
    rangeTSub = rangeData(titleSub)
    print("Range of review title subjectivity: ", rangeTSub)

    #sets variable to function that calculate average polarity of review and prints it out
    global avgReviewPol
    avgReviewPol = calcAvg(reviewPol)
    print("Average review polarity: ", avgReviewPol)

    #sets variable to funtion that finds the standard deviation of Review Polarity
    global stddevRPol
    stddevRPol = standardDev(reviewPol)
    #prints out standard deviation 
    print("Standard Deviation of review polarity: ", stddevRPol)

    #sets variable to function that finds the minimum polarity of review and prints it out 
    global minRPol
    minRPol = minData(reviewPol)
    print("Minimum review polarity: ", minRPol)

    #sets variable to function that finds the maximum polarity of review and prints it out 
    global maxRPol
    maxRPol = maxData(reviewPol)
    print("Maximum review polarity: ", maxRPol)

    #sets variable to function that finds the range of review polarity and prints it out
    global rangeRPol
    rangeRPol = rangeData(reviewPol)
    print("Range of review polarity: ", rangeRPol)
    
    #sets variable to function that calculate the average subjectivity of reviews and prints it out
    global avgReviewSub
    avgReviewSub = calcAvg(reviewSub)
    print("Average review Subjectivity: ", avgReviewSub)

    #sets variable to funtion that finds the standard deviation of Review Subjectivity
    global stddevRSub
    stddevRSub = standardDev(reviewSub)
    #prints out standard deviation 
    print("Standard Deviation of review subjectivity: ", stddevRSub)

    #sets variable to function that finds the minimum subjectivity of reviews and prints it out
    global minRSub
    minRSub = minData(reviewSub)
    print("Minimum review subjectivity: ", minRSub)

    #sets variable to function that finds the maximum subjectivity of reviews and prints it out
    global maxRSub
    maxRSub = maxData(reviewSub)
    print("Maximum review subjectivity: ", maxRSub)

    #sets variable to function that finds the range of review subjectivity and prints it out
    global rangeRSub
    rangeRSub = rangeData(reviewSub)
    print("Range of review subjectivity: ", rangeRSub)

#checks to see if star list is empty 
if len(starList) != 0:
    print(prodNum)
    #sends all varible to function to recieve calculated data for all products and prints out the number of ratings 
    averageForLists(prodNum, starList, titleList, titleSubList, reviewList, reviewSubList)
    prodExcel(sheetW, prodNum, avgStars, avgTitlePol, avgTitleSub, avgReviewPol, avgReviewSub)
    minCalc(sheetW, prodNum, minTPol, minTSub, minRPol, minRSub)
    maxCalc(sheetW, prodNum, maxTPol, maxTSub, maxRPol, maxRSub)
    stdev(sheetW, prodNum, stddev, stddevTPol, stddevTSub, stddevRPol, stddevRSub)
    print("")

#if there are no ratings then send 0 to function for the star list and print out the number of ratings which will be zero
else:
    print(prodNum)
    averageForLists(prodNum, 0, titleList, titleSubList, reviewList, reviewSubList)
    prodExcel(sheetW, prodNum, 0, avgTitlePol, avgTitleSub, avgReviewPol, avgReviewSub)
    minCalc(sheetW, prodNum, minTPol, minTSub, minRPol, minRSub)
    maxCalc(sheetW, prodNum, maxTPol, maxTSub, maxRPol, maxRSub)
    stdev(sheetW, prodNum, 0, stddevTPol, stddevTSub, stddevRPol, stddevRSub)
    print("")

#the same thing as for all products but for product 1  
if len(starList1) != 0:
    averageForLists(prodNum1, starList1, titleList1, titleSubList1, reviewList1, reviewSubList1)
    prodExcel1(sheetW, prodNum1, avgStars, avgTitlePol, avgTitleSub, avgReviewPol, avgReviewSub)
    minCalc1(sheetW, prodNum1, minTPol, minTSub, minRPol, minRSub)
    maxCalc1(sheetW, prodNum1, maxTPol, maxTSub, maxRPol, maxRSub)
    stdev1(sheetW, prodNum1, stddev, stddevTPol, stddevTSub, stddevRPol, stddevRSub)
    print("")
else:
    averageForLists(prodNum1, 0, titleList1, titleSubList1, reviewList1, reviewSubList1)
    prodExcel1(sheetW, prodNum1, 0, avgTitlePol, avgTitleSub, avgReviewPol, avgReviewSub)
    minCalc1(sheetW, prodNum1, minTPol, minTSub, minRPol, minRSub)
    maxCalc1(sheetW, prodNum1, maxTPol, maxTSub, maxRPol, maxRSub)
    stdev1(sheetW, prodNum1, 0, stddevTPol, stddevTSub, stddevRPol, stddevRSub)
    print("")

#the same thing as for all products but for product 2
if len(starList2) != 0:
    averageForLists(prodNum2, starList2, titleList2, titleSubList2, reviewList2, reviewSubList2)
    prodExcel2(sheetW, prodNum2, avgStars, avgTitlePol, avgTitleSub, avgReviewPol, avgReviewSub)
    minCalc2(sheetW, prodNum2, minTPol, minTSub, minRPol, minRSub)
    maxCalc2(sheetW, prodNum2, maxTPol, maxTSub, maxRPol, maxRSub)
    stdev2(sheetW, prodNum2, stddev, stddevTPol, stddevTSub, stddevRPol, stddevRSub)
    print("")
else:
    averageForLists(prodNum2, 0, titleList2, titleSubList2, reviewList2, reviewSubList2)
    prodExcel2(sheetW, prodNum2, 0, avgTitlePol, avgTitleSub, avgReviewPol, avgReviewSub)
    minCalc2(sheetW, prodNum2, minTPol, minTSub, minRPol, minRSub)
    maxCalc2(sheetW, prodNum2, maxTPol, maxTSub, maxRPol, maxRSub)
    stdev2(sheetW, prodNum2, 0, stddevTPol, stddevTSub, stddevRPol, stddevRSub)
    print("")

#the same thing as for all products but for product 3
if len(starList3) != 0:
    averageForLists(prodNum3, starList3, titleList3, titleSubList3, reviewList3, reviewSubList3)
    prodExcel3(sheetW, prodNum3, avgStars, avgTitlePol, avgTitleSub, avgReviewPol, avgReviewSub)
    minCalc3(sheetW, prodNum3, minTPol, minTSub, minRPol, minRSub)
    maxCalc3(sheetW, prodNum3, maxTPol, maxTSub, maxRPol, maxRSub)
    stdev3(sheetW, prodNum3, stddev, stddevTPol, stddevTSub, stddevRPol, stddevRSub)
    print("")
else:
    averageForLists(prodNum3, 0, titleList3, titleSubList3, reviewList3, reviewSubList3)
    prodExcel3(sheetW, prodNum3, 0, avgTitlePol, avgTitleSub, avgReviewPol, avgReviewSub)
    minCalc3(sheetW, prodNum3, minTPol, minTSub, minRPol, minRSub)
    maxCalc3(sheetW, prodNum3, maxTPol, maxTSub, maxRPol, maxRSub)
    stdev3(sheetW, prodNum3, 0, stddevTPol, stddevTSub, stddevRPol, stddevRSub)
    print("")

#the same thing as for all products but for product 4    
if len(starList4) != 0:
    averageForLists(prodNum4, starList4, titleList4, titleSubList4, reviewList4, reviewSubList4)
    prodExcel4(sheetW, prodNum4, avgStars, avgTitlePol, avgTitleSub, avgReviewPol, avgReviewSub)
    minCalc4(sheetW, prodNum4, minTPol, minTSub, minRPol, minRSub)
    maxCalc4(sheetW, prodNum4, maxTPol, maxTSub, maxRPol, maxRSub)
    stdev4(sheetW, prodNum4, stddev, stddevTPol, stddevTSub, stddevRPol, stddevRSub)
    print("")
else:
    averageForLists(prodNum4, 0, titleList4, titleSubList4, reviewList4, reviewSubList4)
    prodExcel4(sheetW, prodNum4, 0, avgTitlePol, avgTitleSub, avgReviewPol, avgReviewSub)
    minCalc4(sheetW, prodNum4, minTPol, minTSub, minRPol, minRSub)
    maxCalc4(sheetW, prodNum4, maxTPol, maxTSub, maxRPol, maxRSub)
    stdev4(sheetW, prodNum4, 0, stddevTPol, stddevTSub, stddevRPol, stddevRSub)
    print("")

#the same thing as for all products but for product 5
if len(starList5) != 0:
    averageForLists(prodNum5, starList5, titleList5, titleSubList5, reviewList5, reviewSubList5)
    prodExcel5(sheetW, prodNum5, avgStars, avgTitlePol, avgTitleSub, avgReviewPol, avgReviewSub)
    minCalc5(sheetW, prodNum5, minTPol, minTSub, minRPol, minRSub)
    maxCalc5(sheetW, prodNum5, maxTPol, maxTSub, maxRPol, maxRSub)
    stdev5(sheetW, prodNum5, stddev, stddevTPol, stddevTSub, stddevRPol, stddevRSub)
    print("")
else:
    averageForLists(prodNum5, 0, titleList5, titleSubList5, reviewList5, reviewSubList5)
    prodExcel5(sheetW, prodNum5, 0, avgTitlePol, avgTitleSub, avgReviewPol, avgReviewSub)
    minCalc5(sheetW, prodNum5, minTPol, minTSub, minRPol, minRSub)
    maxCalc5(sheetW, prodNum5, maxTPol, maxTSub, maxRPol, maxRSub)
    stdev5(sheetW, prodNum5, 0, stddevTPol, stddevTSub, stddevRPol, stddevRSub)
    print("")

#the same thing as for all products but for product 6   
if len(starList6) != 0:
    averageForLists(prodNum6, starList6, titleList6, titleSubList6, reviewList6, reviewSubList6)
    prodExcel6(sheetW, prodNum6, avgStars, avgTitlePol, avgTitleSub, avgReviewPol, avgReviewSub)
    minCalc6(sheetW, prodNum6, minTPol, minTSub, minRPol, minRSub)
    maxCalc6(sheetW, prodNum6, maxTPol, maxTSub, maxRPol, maxRSub)
    stdev6(sheetW, prodNum6, stddev, stddevTPol, stddevTSub, stddevRPol, stddevRSub)
    print("")
else:
    averageForLists(prodNum6, 0, titleList6, titleSubList6, reviewList6, reviewSubList6)
    prodExcel6(sheetW, prodNum6, 0, avgTitlePol, avgTitleSub, avgReviewPol, avgReviewSub)
    minCalc6(sheetW, prodNum6, minTPol, minTSub, minRPol, minRSub)
    maxCalc6(sheetW, prodNum6, maxTPol, maxTSub, maxRPol, maxRSub)
    stdev6(sheetW, prodNum6, 0, stddevTPol, stddevTSub, stddevRPol, stddevRSub)
    print("")


sheetW['R11'].value = len(reviewList1)
sheetW['R12'].value = len(reviewList2)
sheetW['R13'].value = len(reviewList3)
sheetW['R14'].value = len(reviewList4)
sheetW['R15'].value = len(reviewList5)
sheetW['R16'].value = len(reviewList6)
sheetW['R10'].value = "Number of reviews in list"
sheetW['Q11'].value = prodNum1
sheetW['Q12'].value = prodNum2
sheetW['Q13'].value = prodNum3
sheetW['Q14'].value = prodNum4
sheetW['Q15'].value = prodNum5
sheetW['Q16'].value = prodNum6

sheetW['K47'].value = prodNum
sheetW['L46'].value = "Number of all positive reviews"
sheetW['L47'].value = len(rposP)
sheetW['M46'].value = "Number of all negative reviews"
sheetW['M47'].value = len(rnegP)

sheetW['K53'].value = prodNum
sheetW['L52'].value = "Number of all positive review titles"
sheetW['L53'].value = len(tposP)
sheetW['M52'].value = "Number of all negative review titles"
sheetW['M53'].value = len(tnegP)

sheetW['C47'].value = prodNum1
sheetW['C48'].value = prodNum2
sheetW['C49'].value = prodNum3
sheetW['C50'].value = prodNum4
sheetW['C51'].value = prodNum5
sheetW['C52'].value = prodNum6
sheetW['D46'].value = "Number of positive review titles"
sheetW['D47'].value = len(tposP1)
sheetW['D48'].value = len(tposP2)
sheetW['D49'].value = len(tposP3)
sheetW['D50'].value = len(tposP4)
sheetW['D51'].value = len(tposP5)
sheetW['D52'].value = len(tposP6)
sheetW['E46'].value = "Number of negtive review titles"
sheetW['E47'].value = len(tnegP1)
sheetW['E48'].value = len(tnegP2)
sheetW['E49'].value = len(tnegP3)
sheetW['E50'].value = len(tnegP4)
sheetW['E51'].value = len(tnegP5)
sheetW['E52'].value = len(tnegP6)


sheetW['G47'].value = prodNum1
sheetW['G48'].value = prodNum2
sheetW['G49'].value = prodNum3
sheetW['G50'].value = prodNum4
sheetW['G51'].value = prodNum5
sheetW['G52'].value = prodNum6
sheetW['H46'].value = "Number of positive review"
sheetW['H47'].value = len(rposP1)
sheetW['H48'].value = len(rposP2)
sheetW['H49'].value = len(rposP3)
sheetW['H50'].value = len(rposP4)
sheetW['H51'].value = len(rposP5)
sheetW['H52'].value = len(rposP6)
sheetW['I46'].value = "Number of negtive review"
sheetW['I47'].value = len(rnegP1)
sheetW['I48'].value = len(rnegP2)
sheetW['I49'].value = len(rnegP3)
sheetW['I50'].value = len(rnegP4)
sheetW['I51'].value = len(rnegP5)
sheetW['I52'].value = len(rnegP6)

wordcloud1 = WordCloud(width = 1000, height = 500).generate(myText1)
wordcloud2 = WordCloud(width = 1000, height = 500).generate(myText2)
plt.imshow(wordcloud1)
plt.imshow(wordcloud2)
plt.axis("off")
plt.suptitle(wordcloud2)
plt.show()

wk2.save("RatingSummary.xlsx")



        

    
